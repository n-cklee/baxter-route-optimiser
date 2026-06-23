import io
import html as _html
import streamlit as st

DEPOT_SUBURB = "Old Toongabbie"
DEPOT_POSTCODE = "2146"
DEPOT_ADDRESS = "Old Toongabbie NSW 2146"


def _card(title: str, value: str, subtitle: str, bg: str, text_colour: str = "white") -> str:
    return f"""
    <div style="background:{bg};border-radius:12px;padding:20px 24px;
                box-shadow:0 2px 8px rgba(0,0,0,0.15);text-align:center;height:100%;">
        <div style="font-size:13px;color:{text_colour};opacity:0.85;
                    font-weight:600;letter-spacing:0.5px;text-transform:uppercase;">
            {title}
        </div>
        <div style="font-size:42px;font-weight:700;color:{text_colour};
                    line-height:1.2;margin:8px 0 4px;">
            {value}
        </div>
        <div style="font-size:12px;color:{text_colour};opacity:0.75;">
            {subtitle}
        </div>
    </div>
    """


def _html_table(rows: list, header_bg: str = "#003366") -> str:
    """Render a list of dicts as a styled HTML table (no pyarrow dependency)."""
    if not rows:
        return "<p style='color:#666;font-style:italic'>No data.</p>"
    headers = list(rows[0].keys())
    th = "".join(
        f'<th style="padding:8px 12px;text-align:left;white-space:nowrap">'
        f'{_html.escape(str(h))}</th>'
        for h in headers
    )
    body_parts = []
    for i, row in enumerate(rows):
        bg = "#f8f9fa" if i % 2 == 0 else "white"
        tds = "".join(
            f'<td style="padding:7px 12px;border-bottom:1px solid #eee">'
            f'{_html.escape(str(row.get(h) if row.get(h) is not None else ""))}</td>'
            for h in headers
        )
        body_parts.append(f'<tr style="background:{bg}">{tds}</tr>')
    return (
        f'<div style="overflow-x:auto;margin:8px 0">'
        f'<table style="width:100%;border-collapse:collapse;font-size:13px;">'
        f'<thead><tr style="background:{header_bg};color:white">{th}</tr></thead>'
        f'<tbody>{"".join(body_parts)}</tbody>'
        f'</table></div>'
    )


def render_kpi_row(comparison) -> None:
    """Single-card KPI row: AI Optimised."""
    col, _ = st.columns([1, 2])
    with col:
        st.markdown(
            _card(
                "AI Optimised",
                f"{comparison.ai_vans} vans",
                f"{comparison.ai_billing_hours:.1f} billing hours",
                "#00A3E0",
            ),
            unsafe_allow_html=True,
        )

    st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)


def render_wave_breakdown(comparison) -> None:
    """Expandable per-wave comparison table."""
    with st.expander("Per-wave breakdown", expanded=False):
        rows = []
        for w in comparison.waves:
            rows.append({
                "Wave": w.wave,
                "Manual vans": w.manual_van_count,
                "AI vans": w.ai_van_count,
                "Saved vans": w.saved_vans,
                "Manual billing hrs": f"{w.manual_total_hours:.0f}h",
                "AI billing hrs": f"{w.ai_total_hours:.1f}h",
                "Manual drivers": ", ".join(w.manual_drivers) if w.manual_drivers else "—",
            })
        if rows:
            st.markdown(_html_table(rows), unsafe_allow_html=True)


def render_route_cards(routes_by_wave: dict) -> None:
    """Render one card per van across all waves."""
    for wave_key, routes in sorted(routes_by_wave.items()):
        if not routes:
            continue
        st.markdown(f"### Wave: {wave_key}")
        for route in routes:
            colour = route.colour
            status = route.route_status
            if status == "LATE":
                status_label = "🚨 LATE — must split"
                billing_note = f"Actual {route.total_hours:.1f}h — returns after 6 PM, route must be split"
            elif status == "OPTIMAL":
                status_label = "✅ OPTIMAL"
                billing_note = f"Actual {route.total_hours:.1f}h — 9–12h AND back before 6 PM"
            else:
                status_label = "⏱️ UNDER MINIMUM"
                billing_note = f"Actual {route.total_hours:.1f}h — billed for 9h minimum"

            header_html = f"""
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:8px;">
                <div style="width:16px;height:16px;border-radius:50%;
                            background:{colour};flex-shrink:0;"></div>
                <span style="font-size:16px;font-weight:700;">
                    Van {route.van_id} &nbsp;·&nbsp; {route.stop_count} stops
                    &nbsp;·&nbsp; {status_label}
                </span>
            </div>
            <div style="font-size:12px;color:#666;margin-bottom:12px;">
                Departs {route.departure_time.strftime('%H:%M')} &nbsp;|&nbsp; {billing_note}
            </div>
            """
            with st.container():
                st.markdown(header_html, unsafe_allow_html=True)

                rows = []
                for i, (stop, eta) in enumerate(zip(route.stops, route.eta_list)):
                    rows.append({
                        "Consign #": stop.consign_number or "—",
                        "Stop": i + 1,
                        "Receiver": stop.receiver_name,
                        "Suburb": stop.receiver_suburb,
                        "ETA": eta.strftime("%H:%M"),
                        "Dwell": "20 min",
                        "Notify?": "🔔 Yes" if stop.notification_required else "—",
                    })
                # Return-to-depot row
                if route.return_eta:
                    rows.append({
                        "Consign #": "—",
                        "Stop": "↩",
                        "Receiver": f"🏠 Return to Depot ({DEPOT_SUBURB})",
                        "Suburb": DEPOT_SUBURB,
                        "ETA": route.return_eta.strftime("%H:%M"),
                        "Dwell": "—",
                        "Notify?": "—",
                    })
                if rows:
                    st.markdown(_html_table(rows, header_bg=colour), unsafe_allow_html=True)
                st.markdown("<hr style='margin:16px 0;border-color:#eee'>", unsafe_allow_html=True)


def render_debug_panel(routes_by_wave: dict, wave_times: dict) -> None:
    """Collapsible debug panel: wave processing order, van reuse vs new, return times."""
    with st.expander("🔍 Debug: Wave Processing & Van Pool", expanded=False):
        sorted_waves = sorted(wave_times.items(), key=lambda x: x[1])

        seen_van_ids: set = set()
        debug_rows = []

        for wi, (wave_key, departure) in enumerate(sorted_waves):
            routes = routes_by_wave.get(wave_key, [])
            reused = [r for r in routes if r.van_id in seen_van_ids]
            new_vans = [r for r in routes if r.van_id not in seen_van_ids]
            for r in routes:
                seen_van_ids.add(r.van_id)

            return_summary = ", ".join(
                f"Van {r.van_id}→{r.return_eta.strftime('%H:%M') if r.return_eta else '?'}"
                for r in sorted(routes, key=lambda x: x.van_id)
            ) or "—"

            debug_rows.append({
                "Order": wi + 1,
                "Wave": wave_key,
                "Departs": departure.strftime("%H:%M"),
                "Stops": sum(r.stop_count for r in routes),
                "Vans used": len(routes),
                "Reused van IDs": ", ".join(str(r.van_id) for r in sorted(reused, key=lambda x: x.van_id)) or "—",
                "New van IDs": ", ".join(str(r.van_id) for r in sorted(new_vans, key=lambda x: x.van_id)) or "—",
                "Return ETAs": return_summary,
            })

        if debug_rows:
            st.caption(
                "Reused = van already in fleet (returned from earlier wave). "
                "New = fresh van opened for this wave."
            )
            st.markdown(_html_table(debug_rows), unsafe_allow_html=True)
        else:
            st.info("No wave data to display.")


# ── Runsheet generation ───────────────────────────────────────────────────────

def _build_van_route_map(routes_by_wave: dict) -> dict:
    """Return dict[van_id → sorted list of (wave_key, route)]."""
    van_route_map: dict[int, list] = {}
    for wave_key, routes in sorted(routes_by_wave.items()):
        for route in routes:
            van_route_map.setdefault(route.van_id, []).append((wave_key, route))
    for van_id in van_route_map:
        van_route_map[van_id].sort(key=lambda x: x[1].departure_time)
    return van_route_map


def _populate_runsheet_sheet(ws, van_id: int, wave_route_list: list, date_label: str, depot_address: str) -> None:
    """Write one van's runsheet into an openpyxl worksheet."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    header_blue = "003366"
    light_blue  = "D6E4F7"
    mid_grey    = "F2F2F2"
    border_colour = "CCCCCC"

    thin = Side(style="thin", color=border_colour)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_cell(row, col, value, bold=False, size=11, bg=None, align="left", wrap=False, colour="000000"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, color=colour)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    first_dep = wave_route_list[0][1].departure_time
    total_stops = sum(r.stop_count for _, r in wave_route_list)

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value=f"Van {van_id} — Runsheet")
    c.font = Font(bold=True, size=18, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=header_blue)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Meta rows ────────────────────────────────────────────────────────────
    meta = [
        ("Date:", date_label,            "Departure:", first_dep.strftime("%H:%M")),
        ("Depot:", depot_address,         "Total Stops:", str(total_stops)),
    ]
    for ri, (lbl1, val1, lbl2, val2) in enumerate(meta, start=2):
        ws.merge_cells(f"A{ri}:B{ri}")
        ws.merge_cells(f"C{ri}:D{ri}")
        ws.merge_cells(f"E{ri}:F{ri}")
        ws.merge_cells(f"G{ri}:H{ri}")
        for col, text, bold in [(1, lbl1, True), (3, val1, False), (5, lbl2, True), (7, val2, False)]:
            c2 = ws.cell(row=ri, column=col, value=text)
            c2.font = Font(bold=bold, size=11)
            c2.fill = PatternFill("solid", fgColor=light_blue)
            c2.alignment = Alignment(horizontal="left" if not bold else "right", vertical="center")
            c2.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[ri].height = 18

    # ── Blank separator ───────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 6

    # ── Column headers ────────────────────────────────────────────────────────
    HEADERS = ["Stop #", "Consign #", "Receiver Name", "Address (Suburb + Postcode)", "ETA", "Dwell", "Notify?", "Signature"]
    for col, h in enumerate(HEADERS, 1):
        c3 = ws.cell(row=5, column=col, value=h)
        c3.font = Font(bold=True, size=11, color="FFFFFF")
        c3.fill = PatternFill("solid", fgColor=header_blue)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c3.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[5].height = 20

    # ── Stop rows ─────────────────────────────────────────────────────────────
    row_idx = 6
    stop_num = 1
    for wave_key, route in wave_route_list:
        for stop, eta in zip(route.stops, route.eta_list):
            bg = mid_grey if stop_num % 2 == 0 else "FFFFFF"
            address = f"{stop.receiver_suburb} {stop.receiver_postcode}".strip()
            data = [
                stop_num,
                stop.consign_number or "—",
                stop.receiver_name,
                address or "—",
                eta.strftime("%H:%M"),
                "20 min",
                "Yes" if stop.notification_required else "",
                "",  # Signature — blank
            ]
            for col, val in enumerate(data, 1):
                c4 = ws.cell(row=row_idx, column=col, value=val)
                c4.font = Font(size=11)
                c4.fill = PatternFill("solid", fgColor=bg)
                c4.alignment = Alignment(horizontal="center" if col in (1, 5, 6, 7) else "left", vertical="center")
                c4.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.row_dimensions[row_idx].height = 22
            row_idx += 1
            stop_num += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [8, 14, 26, 32, 8, 10, 10, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes below header row
    ws.freeze_panes = "A6"


def _generate_runsheet_bytes(van_route_map: dict, van_ids: list, date_label: str, depot_address: str) -> bytes:
    """Return Excel bytes for the given van_ids from van_route_map."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet
    for van_id in sorted(van_ids):
        wave_route_list = van_route_map[van_id]
        ws = wb.create_sheet(title=f"Van {van_id}")
        _populate_runsheet_sheet(ws, van_id, wave_route_list, date_label, depot_address)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_van_assignments(
    routes_by_wave: dict,
    df_valid,
    date_label: str = "",
    depot_address: str = DEPOT_ADDRESS,
) -> None:
    """
    Van Assignments section: summary table + per-van expandable detail
    with individual runsheet download buttons.
    """
    all_route_pairs = [
        (wave_key, route)
        for wave_key, routes in sorted(routes_by_wave.items())
        for route in routes
    ]

    if not all_route_pairs:
        st.info("No van assignments to display.")
        return

    van_route_map = _build_van_route_map(routes_by_wave)

    # ── Summary table ────────────────────────────────────────────────────────
    summary_rows = []
    for van_id in sorted(van_route_map.keys()):
        wave_route_list = van_route_map[van_id]
        first_dep = wave_route_list[0][1].departure_time
        ret_etas = [r.return_eta for _, r in wave_route_list if r.return_eta]
        last_ret = max(ret_etas) if ret_etas else None
        total_stops = sum(r.stop_count for _, r in wave_route_list)
        total_dur_s = sum(r.total_seconds for _, r in wave_route_list)
        waves_str = ", ".join(w for w, _ in wave_route_list)

        statuses = [r.route_status for _, r in wave_route_list]
        if "LATE" in statuses:
            status = "🚨 LATE — must split"
        elif "UNDER MINIMUM" in statuses:
            status = "⏱️ UNDER MINIMUM"
        else:
            status = "✅ OPTIMAL"

        summary_rows.append({
            "Van #": van_id,
            "First Departure": first_dep.strftime("%H:%M"),
            "Last Return": last_ret.strftime("%H:%M") if last_ret else "—",
            "Total Stops": total_stops,
            "Route Duration": f"{total_dur_s / 3600:.1f}h",
            "Waves Covered": waves_str,
            "Status": status,
        })

    st.markdown("#### Fleet Summary")
    st.markdown(_html_table(summary_rows), unsafe_allow_html=True)
    st.markdown("<div style='margin-top:20px'></div>", unsafe_allow_html=True)

    # ── Per-van expandable detail ─────────────────────────────────────────────
    for van_id in sorted(van_route_map.keys()):
        wave_route_list = van_route_map[van_id]
        total_stops = sum(r.stop_count for _, r in wave_route_list)
        total_dur_s = sum(r.total_seconds for _, r in wave_route_list)
        colour = wave_route_list[0][1].colour
        waves_str = ", ".join(w for w, _ in wave_route_list)

        statuses = [r.route_status for _, r in wave_route_list]
        if "LATE" in statuses:
            status_tag = "🚨 LATE — must split"
        elif "UNDER MINIMUM" in statuses:
            status_tag = "⏱️ UNDER MINIMUM"
        else:
            status_tag = "✅ OPTIMAL"

        expander_label = (
            f"Van {van_id}  ·  Waves: {waves_str}  ·  "
            f"{total_stops} stops  ·  {total_dur_s / 3600:.1f}h total  ·  {status_tag}"
        )

        with st.expander(expander_label, expanded=False):
            for wave_key, route in wave_route_list:
                st.markdown(
                    f"**Wave: {wave_key}** — departs {route.departure_time.strftime('%H:%M')}, "
                    f"returns ~{route.return_eta.strftime('%H:%M') if route.return_eta else '?'}"
                )

                stop_rows = []
                for i, (stop, eta) in enumerate(zip(route.stops, route.eta_list)):
                    stop_rows.append({
                        "Stop #": i + 1,
                        "Consign #": stop.consign_number or "—",
                        "Receiver Name": stop.receiver_name,
                        "Suburb": stop.receiver_suburb,
                        "Postcode": stop.receiver_postcode or "—",
                        "Booking Time": stop.booking_dt.strftime("%H:%M"),
                        "ETA": eta.strftime("%H:%M"),
                        "Dwell": "20 min",
                        "Notification": "🔔 Yes" if stop.notification_required else "—",
                    })

                if route.return_eta:
                    stop_rows.append({
                        "Stop #": "↩",
                        "Consign #": "—",
                        "Receiver Name": "🏠 Return to Depot",
                        "Suburb": DEPOT_SUBURB,
                        "Postcode": DEPOT_POSTCODE,
                        "Booking Time": "—",
                        "ETA": route.return_eta.strftime("%H:%M"),
                        "Dwell": "—",
                        "Notification": "—",
                    })

                st.markdown(_html_table(stop_rows, header_bg=colour), unsafe_allow_html=True)
                st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)

            # Individual van download button
            single_bytes = _generate_runsheet_bytes(van_route_map, [van_id], date_label, depot_address)
            st.download_button(
                label=f"📄 Download Van {van_id} Runsheet",
                data=single_bytes,
                file_name=f"runsheet_van{van_id}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_van_{van_id}",
            )

    # ── All-vans download button ──────────────────────────────────────────────
    st.markdown("<div style='margin-top:24px'></div>", unsafe_allow_html=True)
    all_bytes = _generate_runsheet_bytes(van_route_map, list(van_route_map.keys()), date_label, depot_address)
    st.download_button(
        label="📄 Download Runsheets (All Vans)",
        data=all_bytes,
        file_name="runsheets_all_vans.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_all_vans",
        type="primary",
    )
